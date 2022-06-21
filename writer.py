import os
import sys
from openpyxl import load_workbook, workbook

MONTH = {
    1: 'JANEIRO',
    2: 'FEVEREIRO',
    3: 'MARÇO',
    4: 'ABRIL',
    5: 'MAIO',
    6: 'JUNHO',
    7: 'JULHO',
    8: 'AGOSTO',
    9: 'SETEMBRO',
    10: 'OUTUBRO',
    11: 'NOVEMBRO',
    12: 'DEZEMBRO',
}

def write_workook(data_list, filename):
    if os.path.exists(filename):
        rmfile = input("File already exists, exclude? [Y/n] ")
        if rmfile.lower() != 'y':
            print('ok bye')
            sys.exit(0)
        os.unlink(filename)

    WB = workbook.workbook.Workbook()
    for line in data_list:
        
        cur_month = MONTH.get(line['date'].month)
        
        if not cur_month in WB.sheetnames:
            WB.create_sheet(cur_month)
        
        sheet = WB[cur_month]
        col_date = 'A{}'.format(line['date'].day)
        col_eth = 'B{}'.format(line['date'].day)

        sheet[col_date] = line['date'].strftime('%d/%m/%Y')
        sheet[col_eth] = line['eth_value'].replace('.',',')

    WB.save(filename)

if __name__ == '__main__':
    import datetime
    write_workook([{'date': datetime.datetime(2022, 6, 10, 0, 0), 'eth_value': '0.002326128721417655', 'usd_value': '3.5585582405991567'}])